#!/usr/bin/env python3
"""
Extract each sheet from Ranking_ECAT.xlsx into individual CSV files
in the ranquing-ecat directory.

Each sheet contains ranking data split into masculine (left) and feminine (right)
columns. We parse both sides and save as a unified CSV with a 'gender' column.
"""

import csv
import os
import re
import sys
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: apt-get install -y python3-openpyxl", file=sys.stderr)
    sys.exit(1)

SRC = "/persist/uploads/Ranking_ECAT.xlsx"
DST = Path(__file__).parent.resolve()

os.chdir(DST)


def clean_val(v):
    """Convert cell value to a clean string for CSV export."""
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        if isinstance(v, datetime):
            return v.strftime("%d/%m/%Y")
        return v.isoformat()
    if isinstance(v, str):
        # Clean up newlines in names (relay teams)
        return v.replace("\n", " // ")
    return str(v)


def detect_split_index(ws, max_cols):
    """
    Detect where the left block ends and right block begins.
    Scans the first data rows for empty column gaps that separate
    the masculine (left) and feminine (right) blocks.
    
    Strategy: collect all gaps >= 1 column wide, require at least 3
    non-empty columns on both sides, then pick the gap closest to
    the midpoint of the sheet. Falls back to midpoint if no good gap found.
    Returns the first column index of the right block (0-based).
    """
    midpoint = max_cols // 2
    best_gap = None
    best_distance = max_cols

    for row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=False):
        vals = [cell.value for cell in row]

        # Find gaps of empty columns
        empty_runs = []
        in_empty = False
        run_start = -1
        for i, v in enumerate(vals):
            if v is None:
                if not in_empty:
                    run_start = i
                    in_empty = True
            else:
                if in_empty and run_start >= 0:
                    empty_runs.append((run_start, i))
                    in_empty = False
        if in_empty and run_start >= 0:
            empty_runs.append((run_start, max_cols))

        for start, end in empty_runs:
            gap_size = end - start
            if gap_size >= 1:
                # Require at least 3 non-empty columns before and after the gap
                before_count = sum(1 for v in vals[:start] if v is not None)
                after_count = sum(1 for v in vals[end:] if v is not None)
                if before_count >= 3 and after_count >= 3:
                    # Pick the gap closest to the midpoint
                    gap_center = (start + end) // 2
                    distance = abs(gap_center - midpoint)
                    if distance < best_distance:
                        best_distance = distance
                        best_gap = end

    if best_gap is not None:
        return best_gap

    # Fallback: scan for a single empty column surrounded by content on both sides
    for row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=False):
        vals = [cell.value for cell in row]
        for i in range(1, max_cols - 1):
            if vals[i] is None:
                before_count = sum(1 for v in vals[:i] if v is not None)
                after_count = sum(1 for v in vals[i+1:] if v is not None)
                if before_count >= 3 and after_count >= 3:
                    distance = abs(i - midpoint)
                    if distance < best_distance:
                        best_distance = distance
                        best_gap = i + 1  # return index AFTER the empty column

    if best_gap is not None:
        return best_gap

    return midpoint


def parse_side(ws, col_start, col_end, gender_label):
    """
    Parse one side (masculine or feminine) of the sheet.
    Returns list of dicts with keys: marca, nom, any, lloc, data, vent, categoria
    """
    records = []
    headers = []

    # Detect headers from first few rows
    header_rows = []
    for r in range(1, min(8, ws.max_row + 1)):
        row_vals = []
        has_data = False
        for c in range(col_start, min(col_end, ws.max_column + 1)):
            v = ws.cell(row=r, column=c).value
            row_vals.append(v)
            if v is not None:
                has_data = True
        header_rows.append((r, row_vals, has_data))

    # Find the last header row before data starts
    last_header_row = 0
    for r, row_vals, has_data in header_rows:
        text_vals = [str(v).strip().lower() if v else "" for v in row_vals]
        is_header = any(h in " ".join(text_vals) for h in ["marca", "nom", "any", "lloc", "data", "vent", "categoria", "atleta", "atletes", "puesto", "dorsal"])
        # Also: if all non-empty values look like labels (short text, not numbers/dates)
        non_empty = [v for v in text_vals if v]
        if is_header or (non_empty and all(len(v) < 20 for v in non_empty)):
            last_header_row = r

    if last_header_row == 0:
        # No clear header row found, just start from row 1
        last_header_row = 1

    # Determine column mapping from header row
    header_row_vals = []
    for c in range(col_start, min(col_end, ws.max_column + 1)):
        header_row_vals.append(ws.cell(row=last_header_row, column=c).value)

    # Build column index mapping
    col_map = {}  # index -> column name
    for i, h in enumerate(header_row_vals):
        if h is None:
            continue
        h_lower = str(h).strip().lower()
        if "marca" in h_lower or h_lower == "l":
            col_map[i] = "marca"
        elif "nom" in h_lower or "atleta" in h_lower or "atletes" in h_lower:
            col_map[i] = "nom"
        elif h_lower == "any":
            col_map[i] = "any"
        elif "lloc" in h_lower:
            col_map[i] = "lloc"
        elif "data" in h_lower:
            col_map[i] = "data"
        elif "vent" in h_lower:
            col_map[i] = "vent"
        elif "categoria" in h_lower or "cate" in h_lower:
            col_map[i] = "categoria"
        elif "puesto" in h_lower or "dorsal" in h_lower:
            col_map[i] = "posicio"

    # If no headers detected, fallback to positional mapping for this sheet's known layout
    if "marca" not in col_map:
        # Default: col 0 = marca, col 1 = nom, col 2 = any, col 3 = lloc, col 4 = data
        possible = ["marca", "nom", "any", "lloc", "data", "vent", "categoria"]
        for i, p in enumerate(possible):
            if i < len(header_row_vals):
                col_map[i] = p

    # Parse data rows
    for r in range(last_header_row + 1, ws.max_row + 1):
        row_data = {}
        for c in range(col_start, min(col_end, ws.max_column + 1)):
            idx = c - col_start
            v = ws.cell(row=r, column=c).value
            col_name = col_map.get(idx)
            if col_name:
                row_data[col_name] = v

        # Skip empty rows and header-like rows
        has_any = any(v is not None for v in row_data.values())
        if not has_any:
            continue

        # Skip rows that look like section headers (purely text)
        marca_raw = row_data.get("marca")
        if marca_raw is not None and isinstance(marca_raw, str):
            marca_lower = marca_raw.strip().lower()
            if marca_lower in ("marca", "l", "nom", "any", "lloc", "data", "vent"):
                continue

        # Check if there's at least a name
        nom = row_data.get("nom")
        if nom is None or (isinstance(nom, str) and not nom.strip()):
            # Maybe the "any" column has the name? (some sheets shift columns)
            second_col = row_data.get("any")
            if second_col is not None and isinstance(second_col, str) and len(second_col.strip()) > 3:
                row_data["nom"] = second_col
                row_data["any"] = None
            else:
                continue

        record = {
            "marca": clean_val(row_data.get("marca", "")),
            "nom": clean_val(row_data.get("nom", "")),
            "any": clean_val(row_data.get("any", "")),
            "lloc": clean_val(row_data.get("lloc", "")),
            "data": clean_val(row_data.get("data", "")),
            "vent": clean_val(row_data.get("vent", "")),
            "categoria": clean_val(row_data.get("categoria", "")),
            "gender": gender_label,
        }

        # Only include if we have at least a name and marca
        if record["nom"]:
            records.append(record)

    return records


def expand_relay_teams(records, is_relay_sheet=False):
    """
    Expand multi-athlete relay entries into individual rows.
    Detects entries where 'nom' contains separators like ' // ', hyphens,
    or commas, and creates one row per athlete, preserving marca/lloc/data/gender.

    Also handles cases where column misalignment puts names in the 'any' field
    (4x80 quirk: nom field has the marca value, any field has athlete names).

    Also handles '4x60 Benjami' style where groups of consecutive rows
    share the same marca/lloc/data (same team).
    """

    def is_performance(val):
        """Check if a value looks like a time/performance rather than a name."""
        return bool(re.search(r'["\'″\'\"]', str(val))) or bool(re.match(r'^\d+[\:\']', str(val)))

    def split_any_field(any_val, athlete_count):
        """Try to split the 'any' field to match individual athletes.
        Supports ' // ' separator or space-separated year values."""
        if not any_val or athlete_count <= 1:
            return [any_val] * athlete_count
        parts = [p.strip() for p in any_val.split(" // ") if p.strip()]
        if len(parts) == athlete_count:
            return parts
        # Try space-separated years
        parts = [p.strip() for p in any_val.split() if p.strip()]
        if len(parts) == athlete_count:
            return parts
        return [any_val] * athlete_count

    def split_athletes(nom_field):
        """Try multiple separators to split athlete names. Returns (athletes[], separator_used or None)."""
        # 1. ' // ' separator (from \n in cells)
        if " // " in nom_field:
            parts = [a.strip() for a in nom_field.split(" // ") if a.strip()]
            if len(parts) >= 2:
                return parts, " // "

        # 2. Hyphens (when 2+ hyphens suggest a list)
        hyphen_count = nom_field.count("-")
        if hyphen_count >= 2:
            parts = [a.strip() for a in nom_field.split("-") if a.strip()]
            if len(parts) >= 2:
                return parts, "-"

        # 3. Comma+space separator
        if ", " in nom_field:
            parts = [a.strip() for a in nom_field.split(", ") if a.strip()]
            if len(parts) >= 2:
                return parts, ", "

        # 4. Single comma (no space)
        if "," in nom_field:
            parts = [a.strip() for a in nom_field.split(",") if a.strip()]
            if len(parts) >= 2:
                return parts, ","

        return [nom_field], None

    expanded = []
    i = 0
    while i < len(records):
        rec = records[i]
        nom = rec["nom"]

        # For relay sheets: if nom looks like a performance value (not a name),
        # check the 'any' field for athlete names (handles column misalignment)
        name_field = nom
        if is_relay_sheet and is_performance(nom):
            other_field = rec.get("any", "")
            other_has_athletes = (
                " // " in other_field
                or other_field.count("-") >= 2
                or other_field.count(",") >= 2
            )
            if other_field and other_has_athletes:
                # Move the performance value from nom to marca
                if not rec["marca"]:
                    rec["marca"] = nom
                name_field = other_field
                rec["any"] = ""  # Clear any field, will be populated by split

        athletes, sep = split_athletes(name_field)
        if sep is not None and len(athletes) >= 2:
            expanded_any = split_any_field(rec.get("any", ""), len(athletes))
            for idx, athlete in enumerate(athletes):
                row = dict(rec)
                row["nom"] = athlete
                if idx < len(expanded_any):
                    row["any"] = expanded_any[idx]
                expanded.append(row)
            i += 1
            continue

        # For 4x60 Benjami style: consecutive rows with same marca/lloc/data = same team
        if is_relay_sheet and rec["marca"]:
            # Find consecutive rows that share the same marca/lloc/data
            team = [rec]
            j = i + 1
            while j < len(records):
                next_rec = records[j]
                if (next_rec["marca"] == rec["marca"]
                        and next_rec["lloc"] == rec["lloc"]
                        and next_rec["data"] == rec["data"]
                        and next_rec["gender"] == rec["gender"]
                        and next_rec.get("any", "") != "1.0"  # Not a position row
                        and not next_rec.get("vent", "")):
                    team.append(next_rec)
                    j += 1
                else:
                    break

            if len(team) > 1:
                for member in team:
                    row = dict(rec)  # Keep same marca/lloc/data from first row
                    row["nom"] = member["nom"]
                    row["any"] = member.get("any", rec.get("any", ""))
                    row["vent"] = member.get("vent", "")
                    row["categoria"] = member.get("categoria", "")
                    expanded.append(row)
                i = j
                continue

        # Normal single-athlete entry
        expanded.append(rec)
        i += 1

    return expanded


def write_csv(sheet_name, records, is_relay_sheet=False):
    """Write records to a CSV file named after the sheet."""

    # Expand relay teams into individual athletes
    records = expand_relay_teams(records, is_relay_sheet=is_relay_sheet)

    filename = re.sub(r'[\\/*?:"<>| ]', "_", sheet_name.strip())
    filename = re.sub(r'_+', "_", filename)
    filename = filename.strip("_")
    filepath = DST / f"{filename}.csv"

    fieldnames = ["marca", "nom", "any", "lloc", "data", "vent", "categoria", "gender"]

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for rec in records:
            writer.writerow(rec)

    return filepath


def main():
    print(f"Loading {SRC}...")
    wb = openpyxl.load_workbook(SRC, data_only=True)

    sheets = [s for s in wb.sheetnames if s != "REVISIONS"]
    print(f"Sheets to process: {len(sheets)}")

    total_records = 0
    total_files = 0

    for sheet_name in sheets:
        ws = wb[sheet_name]

        if ws.max_row < 2:
            print(f"  {sheet_name}: SKIP (empty)")
            continue

        max_cols = ws.max_column

        # Detect split point between masculine (left) and feminine (right)
        split_0based = detect_split_index(ws, max_cols)
        split_1based = split_0based + 1  # convert 0-based to 1-based for openpyxl

        # Parse left side (masculine) - columns 1 to split_1based-1
        left_records = parse_side(ws, 1, split_1based, "M")
        # Parse right side (feminine) - columns split_1based to max_cols
        right_records = parse_side(ws, split_1based, max_cols + 1, "F")

        records = left_records + right_records

        # Try to detect gender from the header row
        # Some sheets have MASCULÍ / FEMENÍ labels
        for r in range(1, min(5, ws.max_row + 1)):
            for c in range(1, max_cols + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None and isinstance(v, str):
                    v_upper = v.strip().upper()
                    if "MASCUL" in v_upper:
                        side = "M" if c < split_1based else "F"
                        for rec in records:
                            if abs(records.index(rec) - (left_records.index(rec) if rec in left_records else len(left_records) + right_records.index(rec))) < 1000:
                                pass  # handled by parsing

        if not records:
            print(f"  {sheet_name}: SKIP (no records parsed)")
            continue

        # Detect if this is a relay sheet (contains "x" like 4x60, 3x600, 4x100...)
        is_relay = bool(re.search(r'\d+x\d+', sheet_name.replace(" ", "")))

        filepath = write_csv(sheet_name, records, is_relay_sheet=is_relay)
        count = len(records)
        total_records += count
        total_files += 1

        masc_count = sum(1 for r in records if r["gender"] == "M")
        fem_count = sum(1 for r in records if r["gender"] == "F")
        print(f"  {sheet_name}: {count} records ({masc_count}M / {fem_count}F) -> {filepath.name}")

    print(f"\nDone! {total_files} files, {total_records} total records in {DST}")


if __name__ == "__main__":
    main()