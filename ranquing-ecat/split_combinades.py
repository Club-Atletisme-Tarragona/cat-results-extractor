#!/usr/bin/env python3
"""
Split Combinades sheet into one CSV per combined event type.
Handles:
- Left (M) cols B-H, Right (F) cols J-P
- Carry-forward event type when column not repeated in every row
- Accent normalization (Pentatlo/Pentatló → Pentatló, Tetratlo/Tetratló → Tetratló)
"""
import csv
import os
import re
import sys
from collections import defaultdict

sys.path.insert(0, '/usr/lib/python3/dist-packages')
import openpyxl

XLSX = '/persist/uploads/Ranking_ECAT.xlsx'
OUTDIR = os.path.dirname(os.path.abspath(__file__))

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb['Combinades']

HEADER = ['marca', 'nom', 'any', 'lloc', 'data', 'vent', 'categoria', 'prova', 'gender']
COL_LETTERS = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H',
               10: 'J', 11: 'K', 12: 'L', 13: 'M', 14: 'N', 15: 'O', 16: 'P'}

def fmt_date(v):
    if v is None:
        return ''
    if hasattr(v, 'strftime'):
        return v.strftime('%d/%m/%Y')
    return str(v).strip()

def clean_str(v):
    if v is None:
        return ''
    s = str(v).strip()
    return s

def normalize_event(name):
    """Normalize event name: remove parentheses content and unify accents."""
    if not name:
        return ''
    n = name.strip()
    # Remove trailing parenthetical content like "(PC)" but keep the base
    n = re.sub(r'\s*\(.*?\)\s*$', '', n).strip()
    # Normalize accents
    n = n.replace('Pentatlo', 'Pentatló').replace('Tetratlo', 'Tetratló')
    n = n.replace('Hexatlo', 'Hexatló').replace('Heptatlo', 'Heptatló')
    n = n.replace('Octalo', 'Octaló')
    n = n.replace('Triatlo', 'Triatló')
    return n

def extract_side(row_dict, col_base, carry):
    """
    Extract athlete data from one side.
    col_base: 'B' (M) or 'J' (F)
    carry: dict {(cat_lower, gender): event_type} for category-scoped carry-forward
    """
    pts = row_dict.get(col_base)
    name_cell = row_dict.get(chr(ord(col_base) + 1), None)
    
    pts_val = pts.value if pts else None
    name_val = name_cell.value if name_cell else None
    
    if not pts_val and not name_val:
        return None
    
    name_str = clean_str(name_val)
    pts_str = clean_str(pts_val)
    
# Check if pts_str is numeric (handles values like "1257.0", "1239*")
    def is_numeric_pts(s):
        if not s:
            return False
        # Strip trailing * marker
        clean = s.rstrip('*')
        try:
            float(clean)
            return True
        except ValueError:
            return False
    
    # Skip label-only rows (text in "marca" field like "Tetrató (60 llisos...)")
    if pts_str and not is_numeric_pts(pts_str):
        return None
    if not pts_val and not name_str:
        return None
    if pts_val is None or pts_str == '':
        if not name_str:
            return None
        if not any(c.isdigit() for c in name_str):
            return None
        if len(name_str) > 5:
            return None
        return None
    
    gender = 'M' if col_base == 'B' else 'F'
    
    byear_cell = row_dict.get(chr(ord(col_base) + 2))
    venue_cell = row_dict.get(chr(ord(col_base) + 3))
    date_cell = row_dict.get(chr(ord(col_base) + 4))
    cat_cell = row_dict.get(chr(ord(col_base) + 5))
    evt_cell = row_dict.get(chr(ord(col_base) + 6))  # H or P
    
    byear_val = clean_str(byear_cell.value if byear_cell else '')
    venue_val = clean_str(venue_cell.value if venue_cell else '')
    date_val = fmt_date(date_cell.value if date_cell else '')
    cat_val = clean_str(cat_cell.value if cat_cell else '')
    evt_raw = clean_str(evt_cell.value if evt_cell else '')
    
    # Category-scoped carry-forward
    cat_key = (cat_val.lower(), gender) if cat_val else ('', gender)
    
    if evt_raw:
        normalized = normalize_event(evt_raw)
        if normalized:
            carry[cat_key] = normalized
    
    event_type = carry.get(cat_key, '')
    
    # Format points: strip trailing .0, keep star markers
    pts_clean = pts_str.rstrip('0').rstrip('.') if '.' in pts_str else pts_str
    
    # Clean birth year
    byear_clean = str(int(float(byear_val))) if byear_val and byear_val.replace('.0', '').replace('.','').isdigit() else byear_val
    
    row = {
        'marca': pts_clean,
        'nom': name_str,
        'any': byear_clean,
        'lloc': venue_val,
        'data': date_val,
        'vent': '',
        'categoria': cat_val,
        'prova': event_type if event_type else (f'General_{cat_val}' if cat_val else 'General'),
        'gender': gender,
    }
    return row

# Process all rows with carry-forward per gender
events = defaultdict(list)
carry = {}  # (category_lower, gender) -> event_type

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=16, values_only=False):
    row_dict = {}
    for cell in row:
        if cell.column in COL_LETTERS:
            row_dict[COL_LETTERS[cell.column]] = cell
    
    m_row = extract_side(row_dict, 'B', carry)
    f_row = extract_side(row_dict, 'J', carry)
    
    if m_row:
        events[m_row['prova']].append(m_row)
    if f_row:
        events[f_row['prova']].append(f_row)

# Write files
total = 0
written = []
for event_name in sorted(events.keys()):
    if not event_name:
        continue
    safe = event_name.replace(' ', '_').replace('/', '_').replace('(', '').replace(')', '')
    fpath = os.path.join(OUTDIR, f'Combinades_{safe}.csv')
    
    rows = events[event_name]
    rows.sort(key=lambda r: (r['gender'],
                             -float(re.sub(r'[^\d.]', '', r['marca'])) if re.sub(r'[^\d.]', '', r['marca']) and re.sub(r'[^\d.]', '', r['marca']).replace('.','').isdigit() else 0))
    
    with open(fpath, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=HEADER)
        w.writeheader()
        w.writerows(rows)
    
    cnt = len(rows)
    total += cnt
    m_cnt = sum(1 for r in rows if r['gender'] == 'M')
    f_cnt = sum(1 for r in rows if r['gender'] == 'F')
    cats = set(r['categoria'] for r in rows)
    print(f'  {safe}.csv: {cnt} registres ({m_cnt}M / {f_cnt}F) — categories: {", ".join(sorted(cats))}')
    written.append(safe)

print(f'\nTotal: {total} registres en {len(written)} fitxers')

# Clean up old Combinades.csv
old_file = os.path.join(OUTDIR, 'Combinades.csv')
if os.path.exists(old_file):
    os.remove(old_file)
    print(f'\n🗑️  Eliminat {old_file} (ara dividit en fitxers per prova)')